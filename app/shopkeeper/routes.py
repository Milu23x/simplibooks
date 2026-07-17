import os
from datetime import date, datetime
from calendar import monthrange

from flask import (
    Blueprint, render_template, redirect, url_for, flash, request,
    current_app, send_file, abort
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import (
    Business, Invoice, InvoiceItem, Payment, Expense, ExpenseCategory,
    TaxRate, AuditLog, Notification, Product, PreBooking, Complaint
)
from app.decorators import role_required

shopkeeper_bp = Blueprint(
    "shopkeeper", __name__, url_prefix="/shop", template_folder="../templates/shopkeeper"
)


def _log(action):
    db.session.add(AuditLog(user_id=current_user.id, action=action))
    db.session.commit()


def _business():
    biz = Business.query.filter_by(user_id=current_user.id).first()
    if biz is None:
        abort(404)
    return biz


# ---------------------------------------------------------------- dashboard
@shopkeeper_bp.route("/dashboard")
@login_required
@role_required("shopkeeper")
def dashboard():
    biz = _business()
    today = date.today()
    first_of_month = today.replace(day=1)

    month_invoices = Invoice.query.filter(
        Invoice.business_id == biz.id, Invoice.date >= first_of_month
    ).all()
    month_expenses = Expense.query.filter(
        Expense.business_id == biz.id, Expense.date >= first_of_month
    ).all()

    income_this_month = sum(i.total_amount for i in month_invoices)
    expense_this_month = sum(e.amount for e in month_expenses)
    profit_this_month = income_this_month - expense_this_month

    # Last 6 months trend for the chart
    trend_labels, trend_income, trend_expense = [], [], []
    year, month = today.year, today.month
    months_back = []
    for i in range(5, -1, -1):
        m = month - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        months_back.append((y, m))

    for y, m in months_back:
        start = date(y, m, 1)
        end_day = monthrange(y, m)[1]
        end = date(y, m, end_day)
        inc = db.session.query(db.func.coalesce(db.func.sum(Invoice.total_amount), 0)).filter(
            Invoice.business_id == biz.id, Invoice.date >= start, Invoice.date <= end
        ).scalar()
        exp = db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0)).filter(
            Expense.business_id == biz.id, Expense.date >= start, Expense.date <= end
        ).scalar()
        trend_labels.append(start.strftime("%b %Y"))
        trend_income.append(float(inc))
        trend_expense.append(float(exp))

    recent_invoices = Invoice.query.filter_by(business_id=biz.id).order_by(Invoice.created_at.desc()).limit(5).all()
    outstanding = sum(i.balance_due for i in Invoice.query.filter_by(business_id=biz.id).all())

    return render_template(
        "shopkeeper/dashboard.html",
        biz=biz,
        income_this_month=income_this_month,
        expense_this_month=expense_this_month,
        profit_this_month=profit_this_month,
        recent_invoices=recent_invoices,
        outstanding=outstanding,
        trend_labels=trend_labels,
        trend_income=trend_income,
        trend_expense=trend_expense,
    )


# --------------------------------------------------------- announcements
@shopkeeper_bp.route("/announcements")
@login_required
@role_required("shopkeeper")
def announcements():
    rows = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.date.desc()).all()
    return render_template("shopkeeper/announcements.html", announcements=rows)


# ---------------------------------------------------------------- invoices
@shopkeeper_bp.route("/invoices")
@login_required
@role_required("shopkeeper")
def invoices():
    biz = _business()
    q = Invoice.query.filter_by(business_id=biz.id)

    search = request.args.get("q", "").strip()
    if search:
        q = q.filter(Invoice.customer_name.ilike(f"%{search}%"))

    invs = q.order_by(Invoice.created_at.desc()).all()
    return render_template("shopkeeper/invoices.html", invoices=invs, search=search)


@shopkeeper_bp.route("/invoices/new", methods=["GET", "POST"])
@login_required
@role_required("shopkeeper")
def invoice_new():
    biz = _business()
    tax_rates = TaxRate.query.filter_by(is_active=True).all()

    if request.method == "POST":
        customer_name = request.form.get("customer_name", "").strip()
        customer_state = request.form.get("customer_state", "").strip()
        item_names = request.form.getlist("item_name")
        qtys = request.form.getlist("item_qty")
        rates = request.form.getlist("item_rate")
        tax_rates_selected = request.form.getlist("item_tax")

        if not customer_name or not item_names:
            flash("Customer name and at least one item are required.", "danger")
            return render_template("shopkeeper/invoice_form.html", tax_rates=tax_rates)

        count = Invoice.query.filter_by(business_id=biz.id).count() + 1
        invoice_number = f"INV-{datetime.utcnow().year}-{count:04d}"

        invoice = Invoice(
            business_id=biz.id,
            invoice_number=invoice_number,
            customer_name=customer_name,
            customer_state=customer_state,
            date=date.today(),
            status="unpaid",
        )
        db.session.add(invoice)
        db.session.flush()

        subtotal = 0.0
        tax_total = 0.0
        for name, qty, rate, tax_pct in zip(item_names, qtys, rates, tax_rates_selected):
            if not name.strip():
                continue
            qty = float(qty or 0)
            rate = float(rate or 0)
            tax_pct = float(tax_pct or 0)
            line_amount = qty * rate
            line_tax = line_amount * tax_pct / 100

            db.session.add(InvoiceItem(
                invoice_id=invoice.id, item_name=name, qty=qty, rate=rate,
                tax_rate=tax_pct, amount=line_amount,
            ))
            subtotal += line_amount
            tax_total += line_tax

        invoice.subtotal = round(subtotal, 2)
        invoice.tax_amount = round(tax_total, 2)
        invoice.total_amount = round(subtotal + tax_total, 2)
        db.session.commit()

        _log(f"Created invoice {invoice.invoice_number}")
        flash(f"Invoice {invoice.invoice_number} created.", "success")
        return redirect(url_for("shopkeeper.invoice_view", invoice_id=invoice.id))

    return render_template("shopkeeper/invoice_form.html", tax_rates=tax_rates)


@shopkeeper_bp.route("/invoices/<int:invoice_id>")
@login_required
@role_required("shopkeeper")
def invoice_view(invoice_id):
    biz = _business()
    invoice = Invoice.query.filter_by(id=invoice_id, business_id=biz.id).first_or_404()
    return render_template("shopkeeper/invoice_view.html", invoice=invoice, biz=biz, today=date.today().isoformat())


@shopkeeper_bp.route("/invoices/<int:invoice_id>/pay", methods=["POST"])
@login_required
@role_required("shopkeeper")
def invoice_add_payment(invoice_id):
    biz = _business()
    invoice = Invoice.query.filter_by(id=invoice_id, business_id=biz.id).first_or_404()

    amount = float(request.form.get("amount", 0) or 0)
    mode = request.form.get("mode", "cash")
    pay_date_str = request.form.get("date")
    pay_date = datetime.fromisoformat(pay_date_str).date() if pay_date_str else date.today()
    if amount > 0:
        db.session.add(Payment(invoice_id=invoice.id, amount_paid=amount, mode=mode, date=pay_date))
        db.session.commit()
        invoice.status = "paid" if invoice.balance_due <= 0 else "partially_paid"
        db.session.commit()
        flash("Payment recorded.", "success")
    return redirect(url_for("shopkeeper.invoice_view", invoice_id=invoice.id))


def _generate_invoice_pdf_with_reportlab(invoice, biz, pdf_path):
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter
    margin = 50
    y = height - margin

    c.setFont("Helvetica-Bold", 18)
    c.drawString(margin, y, biz.business_name)
    y -= 24

    c.setFont("Helvetica", 10)
    if biz.gstin:
        c.drawString(margin, y, f"GSTIN: {biz.gstin}")
        y -= 14
    if biz.address:
        c.drawString(margin, y, biz.address)
        y -= 14
    if biz.phone:
        c.drawString(margin, y, f"Phone: {biz.phone}")
        y -= 14

    y -= 10
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, y, f"INVOICE #{invoice.invoice_number}")
    c.drawString(width / 2, y, f"Date: {invoice.date.strftime('%d %b %Y')}")
    y -= 18

    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin, y, "Bill To:")
    c.setFont("Helvetica", 10)
    c.drawString(margin + 60, y, invoice.customer_name)
    y -= 14
    if invoice.customer_state:
        c.drawString(margin + 60, y, f"State: {invoice.customer_state}")
        y -= 14

    y -= 10
    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin, y, "Item")
    c.drawString(margin + 190, y, "Qty")
    c.drawString(margin + 250, y, "Rate")
    c.drawString(margin + 330, y, "Tax")
    c.drawString(margin + 400, y, "Amount")
    y -= 14
    c.setLineWidth(0.5)
    c.line(margin, y, width - margin, y)
    y -= 14

    c.setFont("Helvetica", 10)
    for item in invoice.items:
        if y < margin + 80:
            c.showPage()
            y = height - margin
        c.drawString(margin, y, item.item_name)
        c.drawRightString(margin + 230, y, str(item.qty))
        c.drawRightString(margin + 320, y, f"₹{item.rate:.2f}")
        c.drawRightString(margin + 390, y, f"{item.tax_rate}%")
        c.drawRightString(width - margin, y, f"₹{item.amount:.2f}")
        y -= 14

    y -= 10
    c.line(margin, y, width - margin, y)
    y -= 18

    c.setFont("Helvetica-Bold", 10)
    totals = [
        ("Subtotal", invoice.subtotal),
        ("Tax", invoice.tax_amount),
        ("Total", invoice.total_amount),
        ("Paid", invoice.amount_paid),
        ("Balance Due", invoice.balance_due),
    ]
    for label, value in totals:
        c.drawString(margin, y, label)
        c.drawRightString(width - margin, y, f"₹{value:.2f}")
        y -= 14

    payments = sorted(invoice.payments, key=lambda p: p.date)
    if payments:
        if y < margin + 100:
            c.showPage()
            y = height - margin
        y -= 10
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin, y, "Payment Details")
        y -= 16
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin, y, "Date")
        c.drawString(margin + 150, y, "Mode")
        c.drawRightString(width - margin, y, "Amount Paid")
        y -= 12
        c.line(margin, y, width - margin, y)
        y -= 14
        c.setFont("Helvetica", 10)
        for p in payments:
            if y < margin + 40:
                c.showPage()
                y = height - margin
            c.drawString(margin, y, p.date.strftime("%d %b %Y"))
            c.drawString(margin + 150, y, p.mode_label)
            c.drawRightString(width - margin, y, f"₹{p.amount_paid:.2f}")
            y -= 14
        y -= 6
        status = "PAID IN FULL" if invoice.balance_due <= 0 else ("PARTIALLY PAID" if invoice.amount_paid > 0 else "UNPAID")
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(width - margin, y, f"Payment Status: {status}")
        y -= 14

    c.setFont("Helvetica-Oblique", 9)
    c.drawString(margin, margin, "Generated via SimpliBooks")
    c.save()


@shopkeeper_bp.route("/invoices/<int:invoice_id>/pdf")
@login_required
@role_required("shopkeeper")
def invoice_pdf(invoice_id):
    biz = _business()
    invoice = Invoice.query.filter_by(id=invoice_id, business_id=biz.id).first_or_404()

    pdf_dir = os.path.join(current_app.instance_path, "generated_pdfs")
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_path = os.path.join(pdf_dir, f"{invoice.invoice_number}.pdf")

    try:
        from weasyprint import HTML
        html_string = render_template("shopkeeper/invoice_pdf.html", invoice=invoice, biz=biz)
        HTML(string=html_string).write_pdf(pdf_path)
    except (ImportError, OSError, Exception) as exc:
        current_app.logger.error("WeasyPrint PDF export failed: %s", exc)
        try:
            _generate_invoice_pdf_with_reportlab(invoice, biz, pdf_path)
        except ImportError:
            current_app.logger.error("ReportLab is not installed", exc_info=True)
            flash(
                "PDF export requires WeasyPrint or ReportLab. Install reportlab or system WeasyPrint dependencies.",
                "danger"
            )
            return redirect(url_for("shopkeeper.invoice_view", invoice_id=invoice.id))
        except Exception:
            current_app.logger.exception("ReportLab PDF generation failed")
            flash(
                "Unable to generate the PDF. Please check that ReportLab is installed and invoice data is valid.",
                "danger"
            )
            return redirect(url_for("shopkeeper.invoice_view", invoice_id=invoice.id))

    return send_file(pdf_path, as_attachment=True, download_name=f"{invoice.invoice_number}.pdf")


# ---------------------------------------------------------------- expenses

def _allowed_receipt(filename):
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in current_app.config["ALLOWED_RECEIPT_EXTENSIONS"]


@shopkeeper_bp.route("/expenses")
@login_required
@role_required("shopkeeper")
def expenses():
    biz = _business()
    exps = Expense.query.filter_by(business_id=biz.id).order_by(Expense.date.desc()).all()
    categories = ExpenseCategory.query.all()
    return render_template("shopkeeper/expenses.html", expenses=exps, categories=categories)


@shopkeeper_bp.route("/expenses/new", methods=["POST"])
@login_required
@role_required("shopkeeper")
def expense_new():
    biz = _business()
    amount = float(request.form.get("amount", 0) or 0)
    category_id = request.form.get("category_id") or None
    notes = request.form.get("notes", "").strip()
    exp_date = request.form.get("date") or date.today().isoformat()

    if amount <= 0:
        flash("Expense amount must be greater than zero.", "danger")
        return redirect(url_for("shopkeeper.expenses"))

    receipt_filename = None
    file = request.files.get("receipt")
    if file and file.filename:
        if not _allowed_receipt(file.filename):
            flash("Receipt must be a PNG, JPG, or PDF file.", "danger")
            return redirect(url_for("shopkeeper.expenses"))
        safe_name = secure_filename(f"biz{biz.id}_{datetime.utcnow().timestamp()}_{file.filename}")
        file.save(os.path.join(current_app.config["UPLOAD_FOLDER"], safe_name))
        receipt_filename = safe_name

    db.session.add(Expense(
        business_id=biz.id, category_id=category_id, amount=amount,
        date=datetime.fromisoformat(exp_date).date(), notes=notes, receipt_image=receipt_filename,
    ))
    db.session.commit()
    _log(f"Added expense of {amount}")
    flash("Expense recorded.", "success")
    return redirect(url_for("shopkeeper.expenses"))


@shopkeeper_bp.route("/expenses/<int:expense_id>/delete", methods=["POST"])
@login_required
@role_required("shopkeeper")
def expense_delete(expense_id):
    biz = _business()
    exp = Expense.query.filter_by(id=expense_id, business_id=biz.id).first_or_404()
    db.session.delete(exp)
    db.session.commit()
    flash("Expense deleted.", "info")
    return redirect(url_for("shopkeeper.expenses"))


# ---------------------------------------------------------------- reports
@shopkeeper_bp.route("/reports")
@login_required
@role_required("shopkeeper")
def reports():
    biz = _business()

    start_str = request.args.get("start")
    end_str = request.args.get("end")
    today = date.today()
    start = datetime.fromisoformat(start_str).date() if start_str else today.replace(day=1)
    end = datetime.fromisoformat(end_str).date() if end_str else today

    invs = Invoice.query.filter(Invoice.business_id == biz.id, Invoice.date >= start, Invoice.date <= end).all()
    exps = Expense.query.filter(Expense.business_id == biz.id, Expense.date >= start, Expense.date <= end).all()

    total_income = sum(i.total_amount for i in invs)
    total_tax_collected = sum(i.tax_amount for i in invs)
    total_expense = sum(e.amount for e in exps)
    profit = total_income - total_expense

    # Simple estimated tax liability: flat illustrative rate on net profit.
    # NOTE: this is a rough estimate for planning purposes only, not a filing.
    estimated_tax_rate = 0.30 if profit > 0 else 0.0
    estimated_tax_liability = round(max(profit, 0) * estimated_tax_rate, 2)

    return render_template(
        "shopkeeper/reports.html", start=start, end=end, invoices=invs, expenses=exps,
        total_income=total_income, total_expense=total_expense, profit=profit,
        total_tax_collected=total_tax_collected, estimated_tax_liability=estimated_tax_liability,
        estimated_tax_rate=estimated_tax_rate,
    )


# ---------------------------------------------------------------- promotions (discount banner + products)
@shopkeeper_bp.route("/promotions", methods=["GET", "POST"])
@login_required
@role_required("shopkeeper")
def promotions():
    biz = _business()

    if request.method == "POST":
        biz.banner_enabled = bool(request.form.get("banner_enabled"))
        biz.banner_title = request.form.get("banner_title", "").strip()
        biz.banner_message = request.form.get("banner_message", "").strip()
        biz.banner_discount_percent = float(request.form.get("banner_discount_percent", 0) or 0)
        biz.banner_cta_text = request.form.get("banner_cta_text", "Shop Now").strip() or "Shop Now"
        db.session.commit()
        _log("Updated storefront discount banner")
        flash("Storefront banner updated.", "success")
        return redirect(url_for("shopkeeper.promotions"))

    products = Product.query.filter_by(business_id=biz.id).order_by(Product.created_at.desc()).all()
    return render_template("shopkeeper/promotions.html", biz=biz, products=products)


@shopkeeper_bp.route("/promotions/products/new", methods=["POST"])
@login_required
@role_required("shopkeeper")
def product_new():
    biz = _business()
    name = request.form.get("name", "").strip()
    if not name:
        flash("Product name is required.", "danger")
        return redirect(url_for("shopkeeper.promotions"))

    db.session.add(Product(
        business_id=biz.id,
        name=name,
        category=request.form.get("category", "").strip() or "General",
        description=request.form.get("description", "").strip(),
        price=float(request.form.get("price", 0) or 0),
        discount_percent=float(request.form.get("discount_percent", 0) or 0),
        image_url=request.form.get("image_url", "").strip(),
        allow_prebooking=bool(request.form.get("allow_prebooking")),
        stock_qty=int(request.form.get("stock_qty", 0) or 0),
        is_active=True,
    ))
    db.session.commit()
    _log(f"Added product '{name}' to storefront")
    flash("Product added to your storefront.", "success")
    return redirect(url_for("shopkeeper.promotions"))


@shopkeeper_bp.route("/promotions/products/<int:product_id>/toggle", methods=["POST"])
@login_required
@role_required("shopkeeper")
def product_toggle(product_id):
    biz = _business()
    product = Product.query.filter_by(id=product_id, business_id=biz.id).first_or_404()
    product.is_active = not product.is_active
    db.session.commit()
    return redirect(url_for("shopkeeper.promotions"))


@shopkeeper_bp.route("/promotions/products/<int:product_id>/delete", methods=["POST"])
@login_required
@role_required("shopkeeper")
def product_delete(product_id):
    biz = _business()
    product = Product.query.filter_by(id=product_id, business_id=biz.id).first_or_404()
    db.session.delete(product)
    db.session.commit()
    flash("Product removed.", "info")
    return redirect(url_for("shopkeeper.promotions"))


# ---------------------------------------------------------------- pre-bookings
@shopkeeper_bp.route("/prebookings")
@login_required
@role_required("shopkeeper")
def prebookings():
    biz = _business()
    rows = (
        PreBooking.query.join(Product).filter(Product.business_id == biz.id)
        .order_by(PreBooking.created_at.asc()).all()
    )
    return render_template("shopkeeper/prebookings.html", prebookings=rows)


@shopkeeper_bp.route("/prebookings/<int:prebooking_id>/fulfill", methods=["POST"])
@login_required
@role_required("shopkeeper")
def prebooking_fulfill(prebooking_id):
    biz = _business()
    pb = PreBooking.query.join(Product).filter(
        PreBooking.id == prebooking_id, Product.business_id == biz.id
    ).first_or_404()
    pb.status = "fulfilled"
    db.session.commit()
    flash(f"Marked {pb.customer_name}'s pre-booking as fulfilled.", "success")
    return redirect(url_for("shopkeeper.prebookings"))


# ---------------------------------------------------------------- complaints / contact
@shopkeeper_bp.route("/complaints")
@login_required
@role_required("shopkeeper")
def complaints():
    biz = _business()
    rows = Complaint.query.filter_by(business_id=biz.id).order_by(Complaint.created_at.desc()).all()
    return render_template("shopkeeper/complaints.html", complaints=rows)


@shopkeeper_bp.route("/reports/export")
@login_required
@role_required("shopkeeper")
def reports_export():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    biz = _business()

    start_str = request.args.get("start")
    end_str = request.args.get("end")
    today = date.today()
    start = datetime.fromisoformat(start_str).date() if start_str else today.replace(day=1)
    end = datetime.fromisoformat(end_str).date() if end_str else today

    invs = (
        Invoice.query.filter(Invoice.business_id == biz.id, Invoice.date >= start, Invoice.date <= end)
        .order_by(Invoice.date).all()
    )
    exps = (
        Expense.query.filter(Expense.business_id == biz.id, Expense.date >= start, Expense.date <= end)
        .order_by(Expense.date).all()
    )

    total_income = sum(i.total_amount for i in invs)
    total_tax_collected = sum(i.tax_amount for i in invs)
    total_expense = sum(e.amount for e in exps)
    profit = total_income - total_expense

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(bold=True, size=14, color="2F5496")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(12, length + 2)

    wb = Workbook()

    # --- Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{biz.business_name} — Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Period: {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
    rows = [
        ("Total Income", total_income),
        ("Total Expenses", total_expense),
        ("Net Profit", profit),
        ("GST Collected on Sales", total_tax_collected),
        ("Invoice Count", len(invs)),
        ("Expense Count", len(exps)),
    ]
    for idx, (label, value) in enumerate(rows, start=4):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
    autosize(ws)

    # --- Invoices sheet
    ws2 = wb.create_sheet("Invoices")
    ws2.append(["Invoice #", "Date", "Customer", "Subtotal", "Tax", "Total", "Paid", "Balance Due", "Status"])
    style_header(ws2)
    for i in invs:
        ws2.append([
            i.invoice_number, i.date.strftime("%d-%b-%Y"), i.customer_name,
            i.subtotal, i.tax_amount, i.total_amount, i.amount_paid, i.balance_due, i.status,
        ])
    autosize(ws2)

    # --- Payments sheet (cash / UPI breakdown with dates)
    ws3 = wb.create_sheet("Payments")
    ws3.append(["Invoice #", "Date", "Mode", "Amount"])
    style_header(ws3)
    for i in invs:
        for p in i.payments:
            ws3.append([i.invoice_number, p.date.strftime("%d-%b-%Y"), p.mode_label, p.amount_paid])
    autosize(ws3)

    # --- Expenses sheet
    ws4 = wb.create_sheet("Expenses")
    ws4.append(["Date", "Category", "Amount", "Notes"])
    style_header(ws4)
    for e in exps:
        ws4.append([e.date.strftime("%d-%b-%Y"), e.category.name if e.category else "-", e.amount, e.notes or ""])
    autosize(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{biz.business_name.replace(' ', '_')}_report_{start.isoformat()}_to_{end.isoformat()}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------- profile
@shopkeeper_bp.route("/profile", methods=["GET", "POST"])
@login_required
@role_required("shopkeeper")
def profile():
    biz = _business()
    if request.method == "POST":
        biz.business_name = request.form.get("business_name", biz.business_name).strip()
        biz.gstin = request.form.get("gstin", "").strip()
        biz.state = request.form.get("state", "").strip()
        biz.address = request.form.get("address", "").strip()
        biz.phone = request.form.get("phone", "").strip()
        db.session.commit()
        flash("Business profile updated.", "success")
        return redirect(url_for("shopkeeper.profile"))
    return render_template("shopkeeper/profile.html", biz=biz)
